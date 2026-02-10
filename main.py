import asyncio
import os
import logging
import random
import aiohttp
import pandas as pd
from dataclasses import dataclass, asdict
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
from tqdm.asyncio import tqdm_asyncio

# --- КОНФИГУРАЦИЯ ---
load_dotenv()

API_KEY = os.getenv("GOOGLE_API_KEY")
API_URL = "https://www.googleapis.com/pagespeedonline/v5/runPagespeed"
INPUT_FILE = 'site.txt'
OUTPUT_FILE = 'audit_report.xlsx'

# НАСТРОЙКИ ПРОИЗВОДИТЕЛЬНОСТИ
NUM_ATTEMPTS = 10             # Проверок на URL
MAX_CONCURRENT_REQUESTS = 3   # Потоков
MAX_RETRIES = 3               # Попыток при ошибке
BASE_TIMEOUT = 60             # Базовый таймаут (увеличивается при ретраях)
DELAY_RANGE = (1.0, 2.0)      # Пауза между запросами

# Логгер
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

if not API_KEY:
    raise ValueError("❌ ОШИБКА: Не найден GOOGLE_API_KEY в .env файле.")

# --- ПОРОГОВЫЕ ЗНАЧЕНИЯ ---
THRESHOLDS = {
    'Score': {'good_min': 90, 'avg_min': 50},
    'FCP':  {'good': 1.8, 'poor': 3.0},
    'LCP':  {'good': 2.5, 'poor': 4.0},
    'CLS':  {'good': 0.1, 'poor': 0.25},
    'SI':   {'good': 3.4, 'poor': 5.8},
    'TTFB': {'good': 0.8, 'poor': 1.8},
    'INP':  {'good': 200, 'poor': 500},
    'TBT':  {'good': 200, 'poor': 600}
}

# --- МОДЕЛЬ ДАННЫХ ---
@dataclass
class PageSpeedResult:
    url: str
    device: str
    attempt: int
    score: int
    fcp: float
    lcp: float
    cls: float
    ttfb: float
    si: float
    inp_tbt_value: int
    inp_tbt_source: str
    metric_type_for_color: str

# --- ИНТЕРФЕЙС ---
def ask_devices():
    print("\n📱 ВЫБЕРИТЕ УСТРОЙСТВА ДЛЯ ПРОВЕРКИ:")
    print("   [1] 📱 Только Мобильные")
    print("   [2] 🖥️  Только Десктоп")
    print("   [3] 🚀 Все (Мобильные + Десктоп)")
    
    while True:
        choice = input("\nВведите номер (1-3) и нажмите Enter: ").strip()
        if choice == '1': return ['mobile']
        elif choice == '2': return ['desktop']
        elif choice == '3': return ['mobile', 'desktop']
        else: print("❌ Неверный ввод.")

# --- СЕТЕВОЙ СЛОЙ ---
async def fetch_metrics(session: aiohttp.ClientSession, url: str, device: str, attempt: int) -> Optional[PageSpeedResult]:
    params = {
        'url': url, 'strategy': device, 'key': API_KEY, 'category': 'performance'
    }

    for try_idx in range(MAX_RETRIES):
        # АДАПТИВНЫЙ ТАЙМАУТ: 60 -> 90 -> 120
        current_timeout = BASE_TIMEOUT + (try_idx * 30)
        
        try:
            async with session.get(API_URL, params=params, timeout=aiohttp.ClientTimeout(total=current_timeout)) as resp:
                
                if resp.status == 200:
                    data = await resp.json()
                    
                    if 'error' in data:
                        logger.warning(f"⚠️ API error {url}: {data['error']['message']}")
                        return None

                    lh = data.get('lighthouseResult', {})
                    audits = lh.get('audits', {})
                    
                    if not audits:
                         logger.warning(f"⚠️ Пустой аудит {url}")
                         continue 

                    def get_val(key, div=1):
                        return audits.get(key, {}).get('numericValue', 0) / div

                    score = int(lh.get('categories', {}).get('performance', {}).get('score', 0) * 100)
                    lab_tbt = int(get_val('total-blocking-time'))
                    
                    # INP Логика
                    real_inp = None
                    src = "LAB (TBT)"
                    m_type = "TBT"

                    # 1. URL
                    metrics_url = data.get('loadingExperience', {}).get('metrics', {})
                    if 'INTERACTION_TO_NEXT_PAINT_MS' in metrics_url:
                        real_inp = metrics_url['INTERACTION_TO_NEXT_PAINT_MS'].get('percentile')
                        src = "REAL (URL)"
                        m_type = "INP"
                    
                    # 2. Origin
                    if real_inp is None:
                        metrics_origin = data.get('originLoadingExperience', {}).get('metrics', {})
                        if 'INTERACTION_TO_NEXT_PAINT_MS' in metrics_origin:
                            real_inp = metrics_origin['INTERACTION_TO_NEXT_PAINT_MS'].get('percentile')
                            src = "REAL (Origin)"
                            m_type = "INP"

                    final_val = int(real_inp) if real_inp is not None else lab_tbt

                    return PageSpeedResult(
                        url=url,
                        device=device,
                        attempt=attempt,
                        score=score,
                        fcp=round(get_val('first-contentful-paint', 1000), 2),
                        lcp=round(get_val('largest-contentful-paint', 1000), 2),
                        cls=round(get_val('cumulative-layout-shift'), 3),
                        ttfb=round(get_val('server-response-time', 1000), 2),
                        si=round(get_val('speed-index', 1000), 2),
                        inp_tbt_value=final_val,
                        inp_tbt_source=src,
                        metric_type_for_color=m_type
                    )

                elif resp.status == 429:
                    wait = 10 + (try_idx * 5)
                    logger.warning(f"⏳ Лимит (429) {url}. Ждем {wait}с...")
                    await asyncio.sleep(wait)
                    continue 
                
                elif resp.status >= 500:
                    logger.warning(f"⚠️ Ошибка сервера ({resp.status}). Ретрай...")
                    await asyncio.sleep(5)
                    continue
                
                else:
                    logger.error(f"❌ HTTP {resp.status} для {url}")
                    return None

        except asyncio.TimeoutError:
            logger.warning(f"⏱️ Таймаут ({current_timeout}с) для {url}. Попытка {try_idx+1}/{MAX_RETRIES}")
        except aiohttp.ClientError as e:
            logger.warning(f"🔌 Сеть {url}: {e}. Попытка {try_idx+1}/{MAX_RETRIES}")
        except Exception as e:
            logger.error(f"❌ Ошибка {url}: {e}")
            return None
        
        await asyncio.sleep(2 + (try_idx * 3))

    logger.error(f"❌ Не удалось проверить {url} после {MAX_RETRIES} попыток.")
    return None

async def worker(sem, session, url, device, attempt, pbar):
    async with sem:
        await asyncio.sleep(random.uniform(*DELAY_RANGE))
        res = await fetch_metrics(session, url, device, attempt)
        pbar.update(1)
        return res

async def main_async(sites, devices_list):
    sem = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)
    tasks = []
    
    async with aiohttp.ClientSession() as session:
        total_ops = len(sites) * len(devices_list) * NUM_ATTEMPTS
        
        with tqdm_asyncio(total=total_ops, desc="🚀 Анализ", unit="req") as pbar:
            for site in sites:
                for dev in devices_list:
                    for i in range(1, NUM_ATTEMPTS + 1):
                        tasks.append(worker(sem, session, site, dev, i, pbar))
            
            results = await asyncio.gather(*tasks)
            
    return [r for r in results if r is not None]

# --- ОБРАБОТКА ДАННЫХ ---
def process_and_save(results: List[PageSpeedResult]):
    if not results:
        logger.warning("Нет данных для сохранения.")
        return

    data_dicts = []
    for r in results:
        d = asdict(r)
        d['device'] = 'Мобильный' if r.device == 'mobile' else 'Десктоп'
        data_dicts.append(d)

    df_raw = pd.DataFrame(data_dicts)
    
    cols_map = {
        'url': 'URL', 'device': 'Устройство', 'attempt': 'Попытка', 
        'score': 'Score (0-100)', 'fcp': 'FCP (сек)', 'lcp': 'LCP (сек)', 
        'cls': 'CLS', 'ttfb': 'TTFB (сек)', 'si': 'SI (сек)',
        'inp_tbt_value': 'INP / TBT (мс)', 'inp_tbt_source': 'Источник'
    }
    df_raw = df_raw.rename(columns=cols_map)
    
    group_cols = ['URL', 'Устройство']
    def mode_val(x): return x.mode()[0] if not x.mode().empty else "Н/Д"

    # СРЕДНЕЕ
    agg_mean = {
        'Score (0-100)': 'mean', 'FCP (сек)': 'mean', 'LCP (сек)': 'mean', 'CLS': 'mean', 
        'TTFB (сек)': 'mean', 'SI (сек)': 'mean', 'INP / TBT (мс)': 'mean',
        'Источник': mode_val, 'metric_type_for_color': mode_val, 'Попытка': 'count'
    }
    df_mean = df_raw.groupby(group_cols).agg(agg_mean).reset_index()
    df_mean = df_mean.rename(columns={'Попытка': 'Успешных проверок'})
    for col in ['FCP (сек)', 'LCP (сек)', 'TTFB (сек)', 'SI (сек)']: df_mean[col] = df_mean[col].round(2)
    df_mean['CLS'] = df_mean['CLS'].round(3)
    df_mean['Score (0-100)'] = df_mean['Score (0-100)'].astype(int)
    df_mean['INP / TBT (мс)'] = df_mean['INP / TBT (мс)'].astype(int)

    # МЕДИАНА
    agg_median = agg_mean.copy()
    for key in ['Score (0-100)', 'FCP (сек)', 'LCP (сек)', 'CLS', 'TTFB (сек)', 'SI (сек)', 'INP / TBT (мс)']:
        agg_median[key] = 'median'
    df_median = df_raw.groupby(group_cols).agg(agg_median).reset_index()
    df_median = df_median.rename(columns={'Попытка': 'Успешных проверок'})
    for col in ['FCP (сек)', 'LCP (сек)', 'TTFB (сек)', 'SI (сек)']: df_median[col] = df_median[col].round(2)
    df_median['CLS'] = df_median['CLS'].round(3)
    df_median['Score (0-100)'] = df_median['Score (0-100)'].astype(int)
    df_median['INP / TBT (мс)'] = df_median['INP / TBT (мс)'].astype(int)

    # Сохранение
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_mean.drop(columns=['metric_type_for_color']).to_excel(writer, sheet_name='Средние', index=False)
        df_median.drop(columns=['metric_type_for_color']).to_excel(writer, sheet_name='Медиана', index=False)
        df_raw.drop(columns=['metric_type_for_color']).to_excel(writer, sheet_name='Сырые данные', index=False)

    apply_styles(OUTPUT_FILE, 'Средние', df_mean)
    apply_styles(OUTPUT_FILE, 'Медиана', df_median)
    apply_styles(OUTPUT_FILE, 'Сырые данные', df_raw)

    logger.info(f"✅ Отчет готов: {OUTPUT_FILE}")

def apply_styles(filename, sheet_name, df_data):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    
    fills = {
        'good': PatternFill("solid", start_color="C6EFCE"),
        'avg': PatternFill("solid", start_color="FFEB9C"),
        'bad': PatternFill("solid", start_color="F2DCDB")
    }
    
    header = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    metric_map = {'Score': 'Score (0-100)', 'FCP': 'FCP (сек)', 'LCP': 'LCP (сек)', 'CLS': 'CLS', 'SI': 'SI (сек)', 'TTFB': 'TTFB (сек)'}

    for row_idx, row_data in df_data.iterrows():
        excel_row = row_idx + 2 
        for key_en, key_ru in metric_map.items():
            if key_ru in header:
                val = row_data[key_ru]
                fill = get_color(key_en, val, fills)
                ws.cell(row=excel_row, column=header[key_ru]).fill = fill
                ws.cell(row=excel_row, column=header[key_ru]).alignment = Alignment(horizontal='center')

        col_inp = 'INP / TBT (мс)'
        if col_inp in header:
            val = row_data[col_inp]
            m_type = row_data.get('metric_type_for_color', 'TBT')
            fill = get_color(m_type, val, fills)
            ws.cell(row=excel_row, column=header[col_inp]).fill = fill
            ws.cell(row=excel_row, column=header[col_inp]).alignment = Alignment(horizontal='center')

        col_src = 'Источник'
        if col_src in header:
            cell = ws.cell(row=excel_row, column=header[col_src])
            if "LAB" in str(cell.value):
                cell.font = Font(italic=True, color="808080")
            else:
                cell.font = Font(bold=True, color="2E7D32")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
    wb.save(filename)

def get_color(metric_name, value, fills):
    rule = THRESHOLDS.get(metric_name)
    if not rule: return fills['bad']
    if metric_name == 'Score':
        if value >= rule['good_min']: return fills['good']
        elif value >= rule['avg_min']: return fills['avg']
        else: return fills['bad']
    if value <= rule['good']: return fills['good']
    elif value < rule['poor']: return fills['avg']
    else: return fills['bad']

def main():
    # Проверка наличия файла
    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ ОШИБКА: Файл '{INPUT_FILE}' не найден в папке со скриптом!")
        print(f"   Создайте файл и добавьте туда список URL (по одному на строке).")
        return

    # Чтение сайтов
    with open(INPUT_FILE, 'r') as f:
        sites = [line.strip() for line in f if line.strip()]

    if not sites:
        print(f"\n⚠️ ВНИМАНИЕ: Файл '{INPUT_FILE}' пуст.")
        return

    # 1. Спрашиваем устройства
    selected_devices = ask_devices()
    
    # 2. Расчет метрик
    total_sites = len(sites)
    devices_count = len(selected_devices)
    total_requests = total_sites * devices_count * NUM_ATTEMPTS

    # 3. Красивый вывод шапки
    print("\n" + "="*60)
    print(f"      🚀 GOOGLE PAGESPEED INSIGHTS BULK ANALYZER")
    print("="*60)
    
    print(f"📂 Входной файл:       {INPUT_FILE}")
    print(f"💾 Файл отчета:        {OUTPUT_FILE}")
    print("-" * 60)
    
    print(f"📊 ПАРАМЕТРЫ ЗАДАЧИ:")
    print(f"   • Сайтов для проверки:   {total_sites}")
    
    # Формируем красивое название устройств
    dev_names = []
    if 'mobile' in selected_devices: dev_names.append("Mobile")
    if 'desktop' in selected_devices: dev_names.append("Desktop")
    print(f"   • Устройства:            {' + '.join(dev_names)}")
    
    print(f"   • Проходов на каждый:    {NUM_ATTEMPTS}")
    print(f"   • Параллельных потоков:  {MAX_CONCURRENT_REQUESTS}")
    print("-" * 60)
    
    print(f"∑  ВСЕГО ЗАПРОСОВ К API:    {total_requests}")
    print("="*60 + "\n")

    print("Запуск анализа... Пожалуйста, не закрывайте окно.")
    
    try:
        results = asyncio.run(main_async(sites, selected_devices))
        
        print("\n" + "="*60)
        if results:
            print(f"✅ Анализ завершен! Обработано запросов: {len(results)}/{total_requests}")
            process_and_save(results)
        else:
            print("❌ Не удалось получить данные ни для одного сайта.")
        print("="*60 + "\n")
        
    except KeyboardInterrupt:
        print("\n\n⛔ Работа скрипта прервана пользователем.")

if __name__ == '__main__':
    main()